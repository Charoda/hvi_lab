"""Workbook-backed dataset helpers for the submission-ready MOSAIC surface."""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .results_manifest import CANONICAL_RESULTS_DIR as ACTIVE_RESULTS_DIR
from .results_manifest import RESULTS_MANIFEST_PATH as ACTIVE_RESULTS_MANIFEST
from .results_manifest import active_results_paths
from .tasks import REPO_ROOT

DATASET_WORKBOOK = REPO_ROOT / "mosaic-bench.xlsx"
RESULTS_SHEET = "Results"


def _sheet_rows(workbook_path: Path, sheet_name: str) -> list[list[str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Workbook sheet {sheet_name!r} not found in {workbook_path}.")
    worksheet = workbook[sheet_name]
    if worksheet.max_row == 0:
        workbook.close()
        return []

    rows: list[list[str]] = []
    for row in worksheet.iter_rows(values_only=True):
        normalized = ["" if value is None else str(value).strip() for value in row]
        last_nonempty = 0
        for idx, value in enumerate(normalized, 1):
            if value:
                last_nonempty = idx
        rows.append(normalized[:last_nonempty] if last_nonempty else [])
    workbook.close()
    return rows


@dataclass(frozen=True)
class DatasetRow:
    chain_id: str
    app: str
    cwe: str
    values: dict[str, str]


def load_dataset_rows(workbook_path: Path | None = None) -> list[DatasetRow]:
    """Load workbook-backed dataset rows from mosaic-bench.xlsx."""
    path = workbook_path or DATASET_WORKBOOK
    rows = _sheet_rows(path, RESULTS_SHEET)
    if not rows:
        return []
    headers = rows[0]
    dataset_rows: list[DatasetRow] = []
    for raw in rows[1:]:
        if not raw:
            continue
        padded = list(raw) + [""] * max(0, len(headers) - len(raw))
        row = {headers[i]: padded[i] for i in range(len(headers)) if headers[i]}
        chain_id = (row.get("Chain ID") or "").strip()
        if not chain_id:
            continue
        dataset_rows.append(
            DatasetRow(
                chain_id=chain_id,
                app=(row.get("App") or "").strip(),
                cwe=(row.get("CWE") or "").strip(),
                values=row,
            )
        )
    return dataset_rows


def load_dataset_chain_ids(workbook_path: Path | None = None) -> list[str]:
    return [row.chain_id for row in load_dataset_rows(workbook_path)]


def load_dataset_chain_set(workbook_path: Path | None = None) -> set[str]:
    return set(load_dataset_chain_ids(workbook_path))


def load_dataset_task_ids(workbook_path: Path | None = None) -> set[str]:
    """Resolve task IDs for workbook-backed chains via the chain registry."""
    from .chain_registry import load_chains

    dataset_ids = load_dataset_chain_set(workbook_path)
    task_ids = {
        chain.task_id
        for chain in load_chains()
        if chain.chain_id in dataset_ids and chain.task_id
    }
    return task_ids


def dataset_registry_gaps(workbook_path: Path | None = None) -> dict[str, list[str]]:
    """Report workbook rows that do not map cleanly onto the live chain registry."""
    from .chain_registry import load_chains

    chain_map = {chain.chain_id: chain for chain in load_chains()}
    missing_registry = sorted(chain_id for chain_id in load_dataset_chain_ids(workbook_path) if chain_id not in chain_map)
    missing_task_id = sorted(
        chain.chain_id
        for chain in chain_map.values()
        if chain.chain_id in load_dataset_chain_set(workbook_path) and not chain.task_id
    )
    missing_poc = sorted(
        chain.chain_id
        for chain in chain_map.values()
        if chain.chain_id in load_dataset_chain_set(workbook_path) and not chain.poc_module
    )
    return {
        "missing_registry": missing_registry,
        "missing_task_id": missing_task_id,
        "missing_poc": missing_poc,
    }


def dataset_summary(workbook_path: Path | None = None) -> dict[str, object]:
    rows = load_dataset_rows(workbook_path)
    return {
        "workbook": str((workbook_path or DATASET_WORKBOOK).resolve()),
        "sheet": RESULTS_SHEET,
        "chain_count": len(rows),
        "apps": sorted({row.app for row in rows if row.app}),
        "cwes": sorted({row.cwe for row in rows if row.cwe}),
        "registry_gaps": dataset_registry_gaps(workbook_path),
    }


def active_results_files(
    results_dir: Path | None = None,
    *,
    manifest_path: Path | None = None,
) -> list[Path]:
    """Return active public JSONL files from benchmark/results/manifest.json."""
    resolved_results_dir = Path(results_dir or ACTIVE_RESULTS_DIR).resolve()
    files = active_results_paths(manifest_path, include_missing=False)
    return sorted(path for path in files if path.parent.resolve() == resolved_results_dir)


def load_active_result_rows(
    results_dir: Path | None = None,
    *,
    manifest_path: Path | None = None,
) -> list[dict[str, object]]:
    """Load parseable rows from the active public benchmark result ledger."""
    rows: list[dict[str, object]] = []
    for path in active_results_files(results_dir, manifest_path=manifest_path):
        for line in path.read_text().splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                payload = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(payload, dict):
                rows.append(payload)
    return rows


def dataset_result_ledger_gaps(
    workbook_path: Path | None = None,
    *,
    results_dir: Path | None = None,
    manifest_path: Path | None = None,
) -> dict[str, object]:
    """Compare workbook chains to the active public results ledger."""
    workbook_chain_ids = set(load_dataset_chain_ids(workbook_path))
    files = active_results_files(results_dir, manifest_path=manifest_path)
    rows = load_active_result_rows(results_dir, manifest_path=manifest_path)
    seen_chain_ids = sorted(
        {
            str(row.get("chain_id", "")).strip()
            for row in rows
            if str(row.get("chain_id", "")).strip()
        }
    )
    seen_chain_set = set(seen_chain_ids)
    return {
        "results_dir": str((results_dir or ACTIVE_RESULTS_DIR).resolve()),
        "manifest_path": str(Path(manifest_path or ACTIVE_RESULTS_MANIFEST).resolve()),
        "active_files": [str(path.resolve()) for path in files],
        "active_file_count": len(files),
        "row_count": len(rows),
        "missing_public_results": sorted(workbook_chain_ids - seen_chain_set),
        "unknown_result_chains": sorted(seen_chain_set - workbook_chain_ids),
    }


def build_dataset_manifest_dict(*, models: Iterable[str] | None = None) -> dict[str, object]:
    return {
        "kind": "mosaic.batch",
        "manifest_version": 1,
        "name": "mosaic-bench",
        "description": "Workbook-backed MOSAIC dataset from mosaic-bench.xlsx",
        "selection": {
            "dataset_workbook": "mosaic-bench.xlsx",
            "sheet": RESULTS_SHEET,
        },
        "execution": {
            "models": list(models or ["codex"]),
            "skip_tested": True,
            "warm": True,
            "output": "benchmark/results/canonical/dataset_codex.jsonl",
        },
        "apps": {
            "tasks_dir": "benchmark/apps",
        },
    }


def manifest_json(*, models: Iterable[str] | None = None) -> str:
    return json.dumps(build_dataset_manifest_dict(models=models), indent=2) + "\n"
